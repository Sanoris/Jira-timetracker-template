import configparser
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font

def get_last_month_range_jql():
    today = datetime.today()
    first_of_this_month = today.replace(day=1)
    last_day_prev_month = first_of_this_month - timedelta(days=2)
    first_day_prev_month = last_day_prev_month.replace(day=1)

    start = first_day_prev_month.strftime("%Y-%m-%d")
    end = first_of_this_month.strftime("%Y-%m-%d")

    return start, end

def extract_ticket_and_summary(issue):
    key = issue.get("key", "")
    summary = issue.get("fields", {}).get("summary", "")
    return key, summary

def fetch_account_id(base_url=None, email=None, api_token=None):
    base_url = config["JIRA"]["base_url"]
    email = config["JIRA"]["email"]
    api_token = config["JIRA"]["api_token"]
    url = f"{base_url}/rest/api/3/myself"
    headers = {"Accept": "application/json"}
    auth = (email, api_token)
    response = requests.get(url, headers=headers, auth=auth)
    response.raise_for_status()
    return response.json().get("accountId", "")

def fetch_issues_from_jira(config_path="jira.config"):
    config = configparser.ConfigParser()
    config.read(config_path)

    base_url = config["JIRA"]["base_url"]
    email = config["JIRA"]["email"]
    api_token = config["JIRA"]["api_token"]
    start, end = get_last_month_range_jql()
    jql = f'assignee was currentUser() AND updated >= "{start}" AND updated <= "{end}" ORDER BY updated DESC'
    # For troubleshooting, try a less restrictive JQL:
    #jql = "ORDER BY updated DESC"
    # print(f"Testing with broad JQL: {jql}")

    url = f"{base_url}/rest/api/3/search"
    headers = {"Accept": "application/json"}
    auth = (email, api_token)
    print(f"Generated JQL: {jql}")
    params = {
        "jql": jql,
        "maxResults": 500,
        "fields": "summary,updated,comment"
    }

    response = requests.get(url, headers=headers, params=params, auth=auth)
    try:
        response.raise_for_status()
    except Exception as e:
        print(f"[!] Jira API call failed: {e}\n\tResponse: {response.text}")
        return [], config
    #print(f"[!] Jira API call success:\n\tResponse: {response.text}")
    data = response.json()
    if not data.get("issues"):
        print(f"[!] Jira API response contains no issues. \n\tRaw response data: {data}")
    return data.get("issues", []), config

def build_timesheet(issues, domain, output_file, start, end):
    entries = []
    for issue in issues:
        key, summary = extract_ticket_and_summary(issue)
        updated = issue.get("fields", {}).get("updated", "")
        comments = issue.get("fields", {}).get("comment", {}).get("comments", [])
        # Track if we've added the first comment yet
        first_comment_added = False
        for idx, comment in enumerate(comments):
            author = comment.get("author", {})
            author_account_id = author.get("accountId", "")
            author_display_name = author.get("displayName", "")
            body = comment.get("body", "")
            created = comment.get("created", "")
            if not created:
                continue
            # Only include comments within the start/end date range
            comment_date = datetime.strptime(created[:10], "%Y-%m-%d").date()
            if comment_date < datetime.strptime(start, "%Y-%m-%d").date() or comment_date > datetime.strptime(end, "%Y-%m-%d").date():
                continue
            date = comment_date.isoformat()
            # Safely handle body as dict (Jira Cloud API v3 returns comment body as dict for rich text)
            if isinstance(body, dict):
                # Try to extract plain text from Atlassian Document Format (ADF)
                def extract_text(adf):
                    if isinstance(adf, str):
                        return adf
                    if isinstance(adf, dict):
                        if adf.get('type') == 'text':
                            return adf.get('text', '')
                        if 'content' in adf:
                            return ''.join([extract_text(c) for c in adf['content']])
                    if isinstance(adf, list):
                        return ''.join([extract_text(c) for c in adf])
                    return ''
                body_text = extract_text(body)
            else:
                body_text = str(body)
            desc = body_text[:60].replace('\n', ' ')
            # Only add the first comment (regardless of user)
            if idx == 0 and not first_comment_added:
                entries.append({
                    "DATE": date,
                    "HOURS": 1.5,
                    "TICKET": key,
                    "DESCRIPTION": summary
                })
                first_comment_added = True
            # Add comments from the current user (API user)
            else:
                entries.append({
                    "DATE": date,
                    "HOURS": 1.5,
                    "TICKET": key,
                    "DESCRIPTION": summary
                })
        # If no comments, fall back to summary/updated
        if not comments:
            if updated:
                update_dt = datetime.strptime(updated[:10], "%Y-%m-%d")
                date = update_dt.date().isoformat()
                entries.append({
                    "DATE": date,
                    "HOURS": 1.5,
                    "TICKET": key,
                    "DESCRIPTION": summary
                })
    df = pd.DataFrame(entries)
    if df.empty or 'DATE' not in df.columns:
        print(f"[!] DataFrame missing 'DATE' column or is empty. Columns: {df.columns.tolist()} | Entries: {len(entries)}")
        return
    #day_counts = df.groupby("DATE")["TICKET"].count().to_dict()
    #df["HOURS"] = df["DATE"].apply(lambda d: round(10 / day_counts[d], 2) if day_counts[d] > 0 else 1.5)
    df_grouped = df.groupby(["DATE", "TICKET", "DESCRIPTION"]).agg({"HOURS": "sum"}).reset_index()
    df_grouped = df_grouped.drop_duplicates(subset=["DATE", "TICKET", "DESCRIPTION"])
    df_grouped["LINK"] = df_grouped["TICKET"].apply(
        lambda key: f'=HYPERLINK("https://{domain}.atlassian.net/browse/{key}", "{key}")' if key else ""
    )
    df_grouped.to_excel(output_file, index=False)
    print("\n[info] [trust:high] DataFrame exported to file. Preview:")
    #print(df_grouped.to_string(index=False, justify='left', col_space=16))
    wb = load_workbook(output_file)
    ws = wb.active
    last_row = ws.max_row + 1
    ws[f"A{last_row}"] = "TOTAL"
    ws[f"B{last_row}"] = f"=SUM(B2:B{last_row - 1})"
    ws[f"A{last_row}"].font = Font(bold=True)
    ws[f"B{last_row}"].font = Font(bold=True)
    wb.save(output_file)
    print(f"Saved timesheet to {output_file}")

if __name__ == "__main__":
    issues, config = fetch_issues_from_jira()
    domain = config["META"].get("domain", "")
    output_file = config["META"].get("output_file", "callahan_timesheet_autopulled.xlsx")
    start, end = get_last_month_range_jql()
    build_timesheet(issues, domain, output_file, start, end)
